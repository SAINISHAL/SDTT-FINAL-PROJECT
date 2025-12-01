"""Excel export utilities."""
import os
import time
import pandas as pd
from file_manager import FileManager
from config import DEPARTMENTS, TARGET_SEMESTERS, PRE_MID, POST_MID
from excel_loader import ExcelLoader
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

class ExcelExporter:
    """Handles exporting of timetables to Excel files."""
    
    def __init__(self, data_frames, schedule_generator):
        self.dfs = data_frames
        self.schedule_gen = schedule_generator
        # Expanded vibrant pastel palette (readable on black text)
        self._palette = [
            "FFCDD2","F8BBD0","E1BEE7","D1C4E9","C5CAE9","BBDEFB","B3E5FC","B2EBF2",
            "B2DFDB","C8E6C9","DCEDC8","F0F4C3","FFF9C4","FFECB3","FFE0B2","FFCCBC",
            "D7CCC8","CFD8DC",
            "F28B82","F7A1C4","B39DDB","9FA8DA","90CAF9","81D4FA","80DEEA","80CBC4",
            "A5D6A7","C5E1A5","E6EE9C","FFF59D","FFE082","FFCC80","FFAB91",
            "AED581","81C784","4DD0E1","4FC3F7","9575CD","F48FB1"
        ]
        # Deterministic color mapping per exported workbook
        self._course_color_map = {}
    
    def _course_from_cell(self, val: str) -> str:
        """Extract a course identifier from a cell value."""
        if val is None:
            return ""
        s = str(val).strip()
        if not s or s == "-" or s.upper() == "FREE" or s.upper().startswith("LUNCH"):
            return ""
        # Common patterns: "CS161", "CS161 (Lab)", "CS161-Lab", "CS161: L"
        # Take up to first space or '(' or ':' or '-'
        for sep in [" (", " -", ":", " "]:
            if sep in s:
                s = s.split(sep)[0]
                break
        return s.strip()
    
    def _color_for_course(self, course: str) -> str:
        """Pick a stable color for the course within the current export."""
        if not course:
            return None
        if course not in self._course_color_map:
            idx = len(self._course_color_map) % len(self._palette)
            self._course_color_map[course] = self._palette[idx]
        return self._course_color_map[course]
    
    def _format_worksheet(self, worksheet, has_index=True, start_row=1):
        """Format worksheet to ensure all text is clearly visible.
        - Auto-adjusts column widths
        - Enables text wrapping
        - Sets appropriate row heights
        - Formats headers (bold, center)
        - Sets alignment for data cells"""
        try:
            # Find the maximum column and row with data
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row == 0 or max_col == 0:
                return
            
            # Format header row
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Identify columns that likely contain course names or long text
            course_name_headers = ['course name', 'name', 'course_name']
            faculty_headers = ['faculty', 'instructor']
            time_slot_headers = ['time', 'slot', '08:00', '09:00', '10:00', '11:00', '12:00', '13:00', 
                                 '14:00', '15:00', '16:00', '17:00', '18:00', '19:00']
            
            # Get header row to identify column types
            header_row = {}
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=start_row, column=col_idx)
                if cell.value is not None:
                    header_value = str(cell.value).lower().strip()
                    header_row[col_idx] = header_value
            
            # First pass: calculate optimal column widths
            column_widths = {}
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                avg_length = 0
                count = 0
                
                # Check all cells in this column
                for row_idx in range(start_row, max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        cell_length = len(cell_value)
                        max_length = max(max_length, cell_length)
                        avg_length += cell_length
                        count += 1
                
                if count > 0:
                    avg_length = avg_length / count
                
                # Determine if this is a course name, faculty, or time slot column
                is_course_name_col = False
                is_faculty_col = False
                is_time_slot_col = False
                header_val = header_row.get(col_idx, '').lower()
                
                for course_name_keyword in course_name_headers:
                    if course_name_keyword in header_val:
                        is_course_name_col = True
                        break
                
                for faculty_keyword in faculty_headers:
                    if faculty_keyword in header_val:
                        is_faculty_col = True
                        break
                
                for time_keyword in time_slot_headers:
                    if time_keyword in header_val or ':' in header_val:
                        is_time_slot_col = True
                        break
                
                # Calculate column width based on content type
                if is_time_slot_col:
                    # Time slots are typically short (e.g., "08:00-08:30")
                    column_widths[col_letter] = max(12, min(max_length * 1.1, 18))
                elif is_course_name_col:
                    # Course names - give very generous width to minimize wrapping
                    # Use much wider columns for better readability
                    if max_length > 50:
                        # Very long course names (e.g., "DESIGN&ANALYSIS OF COMPUTER NETWORKS")
                        # Set width based on content length but ensure minimum of 50 units
                        # Approximate: each character is ~0.8 units for wrapped text
                        column_widths[col_letter] = min(max(50, max_length * 0.75), 60)
                    elif max_length > 30:
                        # Long course names - ensure wide enough to minimize wrapping
                        column_widths[col_letter] = min(max(40, max_length * 0.9), 55)
                    else:
                        # Medium length course names
                        column_widths[col_letter] = min(max(30, max_length * 1.1), 45)
                elif is_faculty_col:
                    # Faculty/Instructor columns - give good width for names
                    if max_length > 40:
                        # Long faculty names with multiple instructors
                        column_widths[col_letter] = min(max(35, max_length * 0.7), 50)
                    elif max_length > 25:
                        # Medium length faculty names
                        column_widths[col_letter] = min(max(30, max_length * 0.9), 45)
                    else:
                        # Short faculty names
                        column_widths[col_letter] = min(max(25, max_length * 1.1), 35)
                elif max_length > 40:
                    # Long text columns (but not course names) - still give good width
                    if max_length > 60:
                        column_widths[col_letter] = min(max(35, max_length * 0.6), 50)
                    else:
                        column_widths[col_letter] = min(max(30, max_length * 0.7), 45)
                else:
                    # Normal text columns
                    if max_length > 30:
                        # Medium length text - use reasonable width for wrapping
                        column_widths[col_letter] = min(max(15, max_length * 0.5), 35)
                    else:
                        # Short text - adjust based on content (1 character ≈ 1.1 units)
                        column_widths[col_letter] = min(max(10, max_length * 1.1), 30)
            
            # Second pass: apply formatting and column widths
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    col_letter = get_column_letter(col_idx)
                    
                    # Format header row
                    if row_idx == start_row:
                        cell.font = header_font
                        cell.alignment = header_alignment
                    else:
                        # Format data cells - wrap text
                        # Left align for course names and faculty, center align for others
                        header_val = header_row.get(col_idx, '').lower()
                        is_course_name = any(keyword in header_val for keyword in course_name_headers)
                        is_faculty = any(keyword in header_val for keyword in faculty_headers)
                        
                        if is_course_name or is_faculty:
                            # Left align course names and faculty for better readability
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        else:
                            # Center align other data
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Apply column width
                    if col_letter in column_widths:
                        worksheet.column_dimensions[col_letter].width = column_widths[col_letter]
            
            # Set row heights for better visibility
            # Header row
            if start_row <= max_row:
                worksheet.row_dimensions[start_row].height = 30
            
            # Data rows - set appropriate height for wrapped text
            for row_idx in range(start_row + 1, max_row + 1):
                # Check if row has any content
                has_content = False
                max_lines = 1
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        has_content = True
                        cell_value = str(cell.value)
                        col_letter = get_column_letter(col_idx)
                        col_width = column_widths.get(col_letter, 12)
                        # Estimate lines needed: approximately 10-12 characters per unit of width
                        chars_per_line = max(col_width * 0.85, 8)  # More conservative estimate
                        lines = max(1, len(cell_value) / chars_per_line)
                        max_lines = max(max_lines, lines)
                
                if has_content:
                    # Set height based on estimated lines (18 units per line for better spacing, minimum 25)
                    # Allow rows to be taller for multi-line content
                    row_height = max(25, min(18 * max_lines + 5, 80))
                    worksheet.row_dimensions[row_idx].height = row_height
            
            # Ensure index column (first column) is wide enough if it exists
            if has_index and max_col > 0:
                worksheet.column_dimensions['A'].width = max(15, worksheet.column_dimensions['A'].width or 15)
                
        except Exception as e:
            print(f"    WARNING: Could not format worksheet: {e}")
    
    def _apply_color_coding(self, worksheet, schedule_df, start_row=1, start_col=1):
        """Apply background colors to timetable cells based on course code."""
        # Build mapping using shared cache to keep colors consistent across sheets
        course_to_color = {}
        for day in schedule_df.index:
            for slot in schedule_df.columns:
                val = schedule_df.loc[day, slot]
                course = self._course_from_cell(val)
                if course and course not in course_to_color:
                    course_to_color[course] = self._color_for_course(course)
        # Apply fills
        # Dataframe written starting at (row=start_row, col=start_col), with header row and index col
        header_rows = 1
        index_cols = 1
        nrows = len(schedule_df.index)
        ncols = len(schedule_df.columns)
        for r in range(nrows):
            for c in range(ncols):
                cell = worksheet.cell(row=start_row + header_rows + r, column=start_col + index_cols + c)
                val = cell.value
                course = self._course_from_cell(val)
                if course and course in course_to_color:
                    color = course_to_color[course]
                    try:
                        cell.fill = PatternFill(fill_type="solid", fgColor=color)
                    except Exception:
                        pass
    
    def _get_course_details_for_session(self, semester, department, session_type):
        """Get course details for a specific department and session.
        Validates that expected courses from division logic match what should be scheduled."""
        try:
            # Get all semester courses
            sem_courses_all = ExcelLoader.get_semester_courses(self.dfs, semester)
            if sem_courses_all.empty:
                return pd.DataFrame()
            
            # Parse LTPSC
            sem_courses_parsed = ExcelLoader.parse_ltpsc(sem_courses_all)
            if sem_courses_parsed.empty:
                return pd.DataFrame()
            
            # Filter for department
            if 'Department' in sem_courses_parsed.columns:
                dept_mask = sem_courses_parsed['Department'].astype(str).str.contains(f"^{department}$", na=False, regex=True)
                dept_courses = sem_courses_parsed[dept_mask].copy()
            else:
                dept_courses = sem_courses_parsed.copy()
            
            if dept_courses.empty:
                return pd.DataFrame()
            
            # Divide by session
            pre_mid_courses, post_mid_courses = ExcelLoader.divide_courses_by_session(dept_courses, department, all_sem_courses=sem_courses_parsed)
            
            # Select appropriate session
            if session_type == PRE_MID:
                session_courses = pre_mid_courses
            else:
                session_courses = post_mid_courses
            
            if session_courses.empty:
                print(f"    WARNING: No courses assigned to {department} {session_type} session")
                return pd.DataFrame()
            
            # Prepare summary data
            summary_columns = ['Course Code', 'Course Name', 'Instructor', 'LTPSC', 'Lectures_Per_Week', 'Tutorials_Per_Week', 'Labs_Per_Week', 'Room Allocated', 'Lab Room Allocated', 'Combined Class']
            available_cols = [col for col in summary_columns if col in session_courses.columns]
            
            summary_df = session_courses[available_cols].copy()
            # Ensure Combined Class column exists
            if 'Combined Class' not in summary_df.columns:
                summary_df['Combined Class'] = 'NO'
            # Ensure Room Allocated columns exist
            if 'Room Allocated' not in summary_df.columns:
                summary_df['Room Allocated'] = ''
            if 'Lab Room Allocated' not in summary_df.columns:
                summary_df['Lab Room Allocated'] = ''
            
            # Format counts as "allocated/required" (e.g., "2/3" means 2 allocated out of 3 required)
            if 'Course Code' in summary_df.columns:
                for idx, row in summary_df.iterrows():
                    course_code = str(row.get('Course Code', '')).strip()
                    if course_code and course_code != 'nan' and course_code:
                        # Get required (expected) counts from course data
                        # Use 0 as default if column doesn't exist or value is missing
                        required_lectures = 0
                        required_tutorials = 0
                        required_labs = 0
                        
                        if 'Lectures_Per_Week' in summary_df.columns:
                            required_lectures = pd.to_numeric(row.get('Lectures_Per_Week', 0), errors='coerce')
                            if pd.isna(required_lectures):
                                required_lectures = 0
                            required_lectures = int(required_lectures)
                        
                        if 'Tutorials_Per_Week' in summary_df.columns:
                            required_tutorials = pd.to_numeric(row.get('Tutorials_Per_Week', 0), errors='coerce')
                            if pd.isna(required_tutorials):
                                required_tutorials = 0
                            required_tutorials = int(required_tutorials)
                        
                        if 'Labs_Per_Week' in summary_df.columns:
                            required_labs = pd.to_numeric(row.get('Labs_Per_Week', 0), errors='coerce')
                            if pd.isna(required_labs):
                                required_labs = 0
                            required_labs = int(required_labs)
                        
                        # Get actual allocated counts from schedule generator
                        actual = self.schedule_gen.get_actual_allocations(semester, department, session_type, course_code)
                        actual_lectures = actual.get('lectures', 0)
                        actual_tutorials = actual.get('tutorials', 0)
                        actual_labs = actual.get('labs', 0)
                        
                        # Get combined class flag from actual allocations (read from input Excel file)
                        combined_used = actual.get('combined_class', False)
                        # Room from schedule_gen allocations
                        room_alloc = actual.get('room', '')
                        lab_room_alloc = actual.get('lab_room', '')
                        # Format as "allocated/required"
                        if 'Lectures_Per_Week' in summary_df.columns:
                            summary_df.at[idx, 'Lectures_Per_Week'] = f"{actual_lectures}/{required_lectures}"
                        if 'Tutorials_Per_Week' in summary_df.columns:
                            summary_df.at[idx, 'Tutorials_Per_Week'] = f"{actual_tutorials}/{required_tutorials}"
                        if 'Labs_Per_Week' in summary_df.columns:
                            summary_df.at[idx, 'Labs_Per_Week'] = f"{actual_labs}/{required_labs}"
                        # Mark combined usage
                        summary_df.at[idx, 'Combined Class'] = 'YES' if combined_used else 'NO'
                        if 'Room Allocated' in summary_df.columns:
                            summary_df.at[idx, 'Room Allocated'] = room_alloc
                        if 'Lab Room Allocated' in summary_df.columns:
                            summary_df.at[idx, 'Lab Room Allocated'] = lab_room_alloc
            
            # Validate: Check if any courses have zero LTPSC (should still be included but may not schedule)
            # Note: Now checking actual allocated values from the formatted strings
            if 'Course Code' in summary_df.columns and 'Lectures_Per_Week' in summary_df.columns:
                # Extract actual values from "allocated/required" format for validation
                actual_lectures_list = []
                actual_tutorials_list = []
                actual_labs_list = []
                
                for idx, row in summary_df.iterrows():
                    # Extract allocated value from "allocated/required" format
                    lec_str = str(row.get('Lectures_Per_Week', '0/0'))
                    tut_str = str(row.get('Tutorials_Per_Week', '0/0'))
                    lab_str = str(row.get('Labs_Per_Week', '0/0'))
                    
                    # Parse "allocated/required" format
                    try:
                        actual_lec = int(lec_str.split('/')[0]) if '/' in lec_str else 0
                        actual_tut = int(tut_str.split('/')[0]) if '/' in tut_str else 0
                        actual_lab = int(lab_str.split('/')[0]) if '/' in lab_str else 0
                    except:
                        actual_lec = 0
                        actual_tut = 0
                        actual_lab = 0
                    
                    actual_lectures_list.append(actual_lec)
                    actual_tutorials_list.append(actual_tut)
                    actual_labs_list.append(actual_lab)
                
                # Check for zero LTPSC
                zero_ltpsc_mask = (
                    (pd.Series(actual_lectures_list) == 0) &
                    (pd.Series(actual_tutorials_list) == 0) &
                    (pd.Series(actual_labs_list) == 0)
                )
                zero_ltpsc = summary_df[zero_ltpsc_mask]
                if not zero_ltpsc.empty:
                    zero_codes = zero_ltpsc['Course Code'].dropna().tolist()
                    print(f"    INFO: {len(zero_codes)} courses with 0-0-0 LTPSC in {department} {session_type}: {', '.join(zero_codes)}")
            
            # Rename columns for better display
            column_rename = {
                'Lectures_Per_Week': 'Lectures/Week',
                'Tutorials_Per_Week': 'Tutorials/Week',
                'Labs_Per_Week': 'Labs/Week',
                'Instructor': 'Faculty'
            }
            summary_df = summary_df.rename(columns=column_rename)
            
            return summary_df
            
        except Exception as e:
            print(f"    WARNING: Could not generate course details: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    
    def export_semester_timetable(self, semester):
        """Export timetable for a specific semester."""
        print(f"\n{'='*60}")
        print(f"GENERATING SEMESTER {semester} TIMETABLE")
        print(f"{'='*60}")
        # Reset color map for each workbook to keep palette consistent within file
        self._course_color_map = {}
        
        filename = f"sem{semester}_timetable.xlsx"
        filepath = FileManager.get_output_path(filename)
        
        # Attempt to open writer, handle PermissionError (file locked by Excel)
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except PermissionError as pe:
            print(f"\nWARNING: Cannot write to {filepath} (Permission denied / file may be open).")
            timestamp = int(time.time())
            alt_filename = f"sem{semester}_timetable_{timestamp}.xlsx"
            alt_filepath = FileManager.get_output_path(alt_filename)
            print(f"Attempting alternative filename: {alt_filename}")
            try:
                writer = pd.ExcelWriter(alt_filepath, engine='openpyxl')
                filepath = alt_filepath
                filename = alt_filename
            except Exception as e:
                print(f"\nFAILED: Could not create {filename}: {e}")
                import traceback
                traceback.print_exc()
                return False
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        # Use the opened writer (writer variable guaranteed) in a context manager
        try:
            with writer as w:
                print(f"Creating {filename}...")
                
                # Write Course_Summary first so workbook always has at least one visible sheet
                try:
                    self._add_course_summary(w, semester)
                except Exception as e:
                    print(f"WARNING: Could not write initial Course_Summary: {e}")
                
                # Generate schedules for each department and session
                department_count = 0
                for department in DEPARTMENTS:
                    print(f"\nProcessing {department}:")
                    
                    # Pre-Mid session
                    print(f"  {PRE_MID} session...")
                    try:
                        pre_mid_schedule = self.schedule_gen.generate_department_schedule(semester, department, PRE_MID)
                    except Exception as e:
                        print(f"    ERROR generating {department} {PRE_MID}: {e}")
                        pre_mid_schedule = self.schedule_gen._initialize_schedule()
                    
                    if pre_mid_schedule is not None:
                        sheet_name = f"{department}_{PRE_MID}"
                        clean_schedule = pre_mid_schedule.replace('Free', '-')
                        
                        # Write schedule first
                        clean_schedule.to_excel(w, sheet_name=sheet_name, index=True, startrow=0)
                        
                        # Apply color coding to schedule grid
                        try:
                            ws = w.sheets[sheet_name]
                            self._apply_color_coding(ws, clean_schedule, start_row=1, start_col=1)
                        except Exception as e:
                            print(f"    WARNING: Could not apply color coding to {sheet_name}: {e}")
                        
                        # Get course details for this session
                        course_details = self._get_course_details_for_session(semester, department, PRE_MID)
                        
                        # Add course details below the schedule
                        if not course_details.empty:
                            # Calculate starting row (schedule rows + header + 2 blank rows)
                            start_row = len(clean_schedule) + 3
                            
                            # Write a header for course details section
                            worksheet = w.sheets[sheet_name]
                            worksheet.cell(row=start_row, column=1, value="COURSE DETAILS:")
                            
                            # Write course details table
                            course_details.to_excel(w, sheet_name=sheet_name, index=False, startrow=start_row+1)
                        
                        # Format worksheet to ensure all text is visible
                        try:
                            ws = w.sheets[sheet_name]
                            self._format_worksheet(ws, has_index=True, start_row=1)
                        except Exception as e:
                            print(f"    WARNING: Could not format {sheet_name}: {e}")
                        
                        print(f"    SUCCESS: {sheet_name} created with course details")
                        department_count += 1
                    else:
                        print(f"    FAILED: {department}_{PRE_MID}")
                    
                    # Post-Mid session  
                    print(f"  {POST_MID} session...")
                    try:
                        post_mid_schedule = self.schedule_gen.generate_department_schedule(semester, department, POST_MID)
                    except Exception as e:
                        print(f"    ERROR generating {department} {POST_MID}: {e}")
                        post_mid_schedule = self.schedule_gen._initialize_schedule()
                    
                    if post_mid_schedule is not None:
                        sheet_name = f"{department}_{POST_MID}"
                        clean_schedule = post_mid_schedule.replace('Free', '-')
                        
                        # Write schedule first
                        clean_schedule.to_excel(w, sheet_name=sheet_name, index=True, startrow=0)
                        
                        # Apply color coding to schedule grid
                        try:
                            ws = w.sheets[sheet_name]
                            self._apply_color_coding(ws, clean_schedule, start_row=1, start_col=1)
                        except Exception as e:
                            print(f"    WARNING: Could not apply color coding to {sheet_name}: {e}")
                        
                        # Get course details for this session
                        course_details = self._get_course_details_for_session(semester, department, POST_MID)
                        
                        # Add course details below the schedule
                        if not course_details.empty:
                            # Calculate starting row (schedule rows + header + 2 blank rows)
                            start_row = len(clean_schedule) + 3
                            
                            # Write a header for course details section
                            worksheet = w.sheets[sheet_name]
                            worksheet.cell(row=start_row, column=1, value="COURSE DETAILS:")
                            
                            # Write course details table
                            course_details.to_excel(w, sheet_name=sheet_name, index=False, startrow=start_row+1)
                        
                        # Format worksheet to ensure all text is visible
                        try:
                            ws = w.sheets[sheet_name]
                            self._format_worksheet(ws, has_index=True, start_row=1)
                        except Exception as e:
                            print(f"    WARNING: Could not format {sheet_name}: {e}")
                        
                        print(f"    SUCCESS: {sheet_name} created with course details")
                        department_count += 1
                    else:
                        print(f"    FAILED: {department}_{POST_MID}")
                
                # Add Electives and Minor sheets
                try:
                    self._add_electives_sheet(w, semester)
                except Exception as e:
                    print(f"WARNING: Could not add Electives sheet: {e}")
                
                try:
                    self._add_minor_sheet(w, semester)
                except Exception as e:
                    print(f"WARNING: Could not add Minor sheet: {e}")
                
                print(f"\nSUCCESS: Created {filename}")
                print(f"  - {department_count} department schedules")
                print(f"  - Course summary sheet")
                print(f"  - Electives sheet")
                print(f"  - Minor sheet")
            
            return True
            
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _add_course_summary(self, writer, semester):
        """Add course information summary. Always create the Course_Summary sheet (may be empty).
        Adds LTPSC validity check for all courses."""
        try:
            # Prepare empty default summary (columns if available)
            default_cols = ['Course Code', 'Course Name', 'LTPSC', 'Credits']
            summary_df = pd.DataFrame(columns=default_cols)

            ltpsc_valid_col = []
            all_valid = True

            if 'course' in self.dfs:
                course_df = self.dfs['course']
                if 'Semester' in course_df.columns:
                    temp_df = course_df.copy()
                    temp_df['Semester'] = pd.to_numeric(temp_df['Semester'], errors='coerce')
                    sem_courses = temp_df[temp_df['Semester'] == semester]

                    if not sem_courses.empty:
                        available_cols = [col for col in default_cols if col in sem_courses.columns]
                        summary_df = sem_courses[available_cols].copy()
                        # Check LTPSC validity for each course
                        for idx, row in summary_df.iterrows():
                            ltpsc_val = str(row.get('LTPSC', '')).strip()
                            valid = False
                            if ltpsc_val and '-' in ltpsc_val:
                                parts = ltpsc_val.split('-')
                                if len(parts) >= 3:
                                    try:
                                        float(parts[0])
                                        float(parts[1])
                                        float(parts[2])
                                        valid = True
                                    except Exception:
                                        valid = False
                            ltpsc_valid_col.append(valid)
                            if not valid:
                                all_valid = False
                        summary_df['LTPSC_Valid'] = ltpsc_valid_col
                        print(f"SUCCESS: Added Course_Summary sheet with {len(summary_df)} courses")
                    else:
                        print(f"WARNING: No courses found for semester {semester}; writing empty Course_Summary")
                else:
                    print("WARNING: 'Semester' column not found in course data; writing empty Course_Summary")
            else:
                print("WARNING: 'course' data frame not found; writing empty Course_Summary")

            # Add a message row at the top
            from pandas import DataFrame
            msg = "All courses follow LTPSC structure." if all_valid and not summary_df.empty else "Some courses do NOT follow LTPSC structure."
            msg_df = DataFrame({'Course Code': [msg]})
            # Write message row, then summary directly to the existing writer
            msg_df.to_excel(writer, sheet_name='Course_Summary', index=False, header=False, startrow=0)
            summary_df.to_excel(writer, sheet_name='Course_Summary', index=False, startrow=2)
            
            # Format Course_Summary worksheet
            try:
                ws = writer.sheets['Course_Summary']
                # Format the message row (row 1) separately
                if ws.max_row > 0:
                    msg_cell = ws.cell(row=1, column=1)
                    msg_cell.font = Font(bold=True, size=11)
                    msg_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    # Merge cells for message if there are multiple columns
                    if ws.max_column > 1:
                        try:
                            ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
                        except:
                            pass
                
                # Format the rest of the sheet (header at row 2, data starts at row 3)
                self._format_worksheet(ws, has_index=False, start_row=2)  # Header is at row 2
            except Exception as e:
                print(f"WARNING: Could not format Course_Summary: {e}")
        except Exception as e:
            print(f"FAILED: Could not add course summary: {e}")
    
    def _assign_room_by_capacity(self, students, semester_id, assigned_rooms=None):
        """Assign rooms for electives/minors based on student strength with deterministic rules."""
        if not self.schedule_gen:
            return ""
        
        try:
            students = int(float(students)) if students else 0
        except (ValueError, TypeError):
            return ""
        
        if students <= 0:
            return ""
        
        if assigned_rooms is None:
            assigned_rooms = set()
        
        nonlab_rooms = list(getattr(self.schedule_gen, 'nonlab_rooms', []))
        room_pool = []
        seen = set()
        for name, cap in nonlab_rooms:
            if not name or name in seen:
                continue
            seen.add(name)
            room_pool.append((name, cap or 0))
        
        c004_room = getattr(self.schedule_gen, 'c004_room', None)
        if c004_room:
            room_pool.append(c004_room)
        
        if not room_pool and getattr(self.schedule_gen, 'classrooms', None):
            seen = set()
            for name, cap in self.schedule_gen.classrooms:
                if not name or name in seen or name in [r[0] for r in getattr(self.schedule_gen, 'lab_rooms', [])]:
                    continue
                seen.add(name)
                room_pool.append((name, cap or 0))
        
        if not room_pool:
            return ""
        
        shared_overlap = set(getattr(self.schedule_gen, 'shared_rooms', set()) or [])
        
        def can_reuse(room_name):
            if not room_name:
                return False
            return room_name.upper() in shared_overlap
        
        def filter_rooms(predicate):
            filtered = []
            for name, cap in room_pool:
                if not name:
                    continue
                if name in assigned_rooms and not can_reuse(name):
                    continue
                if predicate(name.upper(), cap):
                    filtered.append((name, cap))
            return filtered
        
        def pick_smallest(rooms):
            if not rooms:
                return ""
            rooms.sort(key=lambda x: (x[1] if x[1] else float('inf'), x[0]))
            return rooms[0][0]
        
        # 1) Large cohorts (>=120) must use C004 if available
        if students >= 120:
            if c004_room:
                return c004_room[0]
            large_candidates = filter_rooms(lambda _, cap: cap and cap >= students)
            room = pick_smallest(large_candidates) if large_candidates else ""
            if room:
                return room
        
        # 2) Mid cohorts (80-199) prefer C001 / ~120 seat rooms
        if 80 <= students < 200:
            c001_candidate = filter_rooms(lambda name, _: name == 'C001')
            if c001_candidate:
                return c001_candidate[0][0]
            mid_candidates = filter_rooms(lambda _, cap: cap and 80 <= cap <= max(students, 120))
            room = pick_smallest(mid_candidates)
            if room:
                return room
        
        # 3) Smaller cohorts (<80) use smallest normal classroom available
        if students < 80:
            small_candidates = filter_rooms(lambda _, cap: cap and cap >= max(students, 30))
            room = pick_smallest(small_candidates)
            if room:
                return room
        
        # Fallback: choose any remaining non-lab room
        fallback_candidates = filter_rooms(lambda _, cap: cap)
        room = pick_smallest(fallback_candidates)
        if room:
            return room
        
        return ""
    
    def _get_electives_data(self, semester):
        """Get elective data for a specific semester from 'Elective Data' sheet.
        Returns DataFrame with columns: Course Code, Course Name, Faculty, Semester, Students, Classroom"""
        try:
            # Try to find elective data sheet
            elective_df = None
            sheet_keys = [k for k in self.dfs.keys() if 'elective' in k.lower()]
            
            if not sheet_keys:
                # Try loading from course_data.xlsx directly
                try:
                    from config import INPUT_DIR
                    course_file = os.path.join(INPUT_DIR, 'course_data.xlsx')
                    if os.path.exists(course_file):
                        xl_file = pd.ExcelFile(course_file)
                        for sheet_name in xl_file.sheet_names:
                            if 'elective' in sheet_name.lower():
                                elective_df = pd.read_excel(course_file, sheet_name=sheet_name)
                                break
                except Exception as e:
                    pass
            
            if not sheet_keys and elective_df is None:
                # Check if it's in the loaded data frames with different naming
                for key in self.dfs.keys():
                    if 'elective' in key.lower() or 'electives' in key.lower():
                        elective_df = self.dfs[key]
                        break
            
            if sheet_keys and elective_df is None:
                elective_df = self.dfs[sheet_keys[0]]
            
            if elective_df is None or elective_df.empty:
                return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
            
            # Normalize column names - handle uppercase and variations
            elective_df = elective_df.copy()
            column_map = {}
            for col in elective_df.columns:
                col_lower = str(col).strip().lower()
                if any(x in col_lower for x in ['course code', 'code']) and 'name' not in col_lower:
                    column_map[col] = 'Course Code'
                elif any(x in col_lower for x in ['course name', 'coursename']) and 'code' not in col_lower:
                    column_map[col] = 'Course Name'
                elif any(x in col_lower for x in ['faculty', 'instructor', 'teacher']):
                    column_map[col] = 'Faculty'
                elif any(x in col_lower for x in ['semester', 'sem']):
                    column_map[col] = 'Semester'
                elif any(x in col_lower for x in ['student', 'registered', 'enrollment', 'enrol']):
                    column_map[col] = 'Students'
            
            elective_df = elective_df.rename(columns=column_map)
            
            # Filter by semester
            if 'Semester' in elective_df.columns:
                elective_df['Semester'] = pd.to_numeric(elective_df['Semester'], errors='coerce')
                elective_df = elective_df[elective_df['Semester'] == semester].copy()
            
            if elective_df.empty:
                return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
            
            # Prepare output columns
            output_cols = ['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students']
            available_cols = [col for col in output_cols if col in elective_df.columns]
            result_df = elective_df[available_cols].copy()
            
            # Add Classroom column and assign rooms - ensure each course gets a different room
            assigned_rooms = set()
            if 'Students' in result_df.columns:
                def assign_unique_room(students_val):
                    room = self._assign_room_by_capacity(students_val, semester, assigned_rooms)
                    if room:
                        assigned_rooms.add(room)
                    return room
                
                result_df['Classroom'] = result_df['Students'].apply(assign_unique_room)
            else:
                result_df['Classroom'] = ""
            
            # Ensure all required columns exist
            for col in ['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom']:
                if col not in result_df.columns:
                    result_df[col] = ""
            
            # Reorder columns
            result_df = result_df[['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom']]
            
            return result_df
            
        except Exception as e:
            print(f"ERROR: Could not load elective data: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
    
    def _get_minor_data(self, semester):
        """Get minor data for a specific semester from 'Minor Data' sheet.
        Returns DataFrame with columns: Course Code, Course Name, Faculty, Semester, Students, Classroom"""
        try:
            # Try to find minor data sheet
            minor_df = None
            sheet_keys = [k for k in self.dfs.keys() if 'minor' in k.lower()]
            
            if not sheet_keys:
                # Try loading from course_data.xlsx directly
                try:
                    from config import INPUT_DIR
                    course_file = os.path.join(INPUT_DIR, 'course_data.xlsx')
                    if os.path.exists(course_file):
                        xl_file = pd.ExcelFile(course_file)
                        for sheet_name in xl_file.sheet_names:
                            if 'minor' in sheet_name.lower():
                                minor_df = pd.read_excel(course_file, sheet_name=sheet_name)
                                break
                except Exception as e:
                    pass
            
            if not sheet_keys and minor_df is None:
                # Check if it's in the loaded data frames with different naming
                for key in self.dfs.keys():
                    if 'minor' in key.lower() or 'minors' in key.lower():
                        minor_df = self.dfs[key]
                        break
            
            if sheet_keys and minor_df is None:
                minor_df = self.dfs[sheet_keys[0]]
            
            if minor_df is None or minor_df.empty:
                return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
            
            # Normalize column names - handle uppercase and variations
            # Minor data has different structure: 'MINOR COURSE ' instead of separate Code/Name
            minor_df = minor_df.copy()
            column_map = {}
            for col in minor_df.columns:
                col_lower = str(col).strip().lower()
                # Handle "MINOR COURSE " - use as both Course Code and Course Name
                if any(x in col_lower for x in ['minor course', 'course']) and 'semester' not in col_lower and 'student' not in col_lower:
                    column_map[col] = 'Course Name'  # Use as course name, we'll extract code if needed
                elif any(x in col_lower for x in ['faculty', 'instructor', 'teacher']):
                    column_map[col] = 'Faculty'
                elif any(x in col_lower for x in ['semester', 'sem']):
                    column_map[col] = 'Semester'
                elif any(x in col_lower for x in ['student', 'registered', 'enrollment', 'enrol']):
                    column_map[col] = 'Students'
            
            minor_df = minor_df.rename(columns=column_map)
            
            # Filter by semester
            if 'Semester' in minor_df.columns:
                minor_df['Semester'] = pd.to_numeric(minor_df['Semester'], errors='coerce')
                minor_df = minor_df[minor_df['Semester'] == semester].copy()
            
            if minor_df.empty:
                return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
            
            # Prepare output columns
            output_cols = ['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students']
            available_cols = [col for col in output_cols if col in minor_df.columns]
            result_df = minor_df[available_cols].copy()
            
            # For Minor data, if we have Course Name but not Course Code, use Course Name as Code
            # (Minor courses often don't have separate codes)
            if 'Course Name' in result_df.columns and 'Course Code' not in result_df.columns:
                result_df['Course Code'] = result_df['Course Name'].apply(lambda x: str(x).strip().upper()[:8] if x and str(x).strip() != '' and str(x).lower() != 'nan' else "")
            
            # If Course Code exists but Course Name doesn't, use Code as Name
            if 'Course Code' in result_df.columns and 'Course Name' not in result_df.columns:
                result_df['Course Name'] = result_df['Course Code']
            
            # Add Classroom column and assign rooms - ensure each course gets a different room
            assigned_rooms = set()
            if 'Students' in result_df.columns:
                def assign_unique_room(students_val):
                    room = self._assign_room_by_capacity(students_val, semester, assigned_rooms)
                    if room:
                        assigned_rooms.add(room)
                    return room
                
                result_df['Classroom'] = result_df['Students'].apply(assign_unique_room)
            else:
                result_df['Classroom'] = ""
            
            # Ensure all required columns exist
            for col in ['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom']:
                if col not in result_df.columns:
                    result_df[col] = ""
            
            # Reorder columns
            result_df = result_df[['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom']]
            
            return result_df
            
        except Exception as e:
            print(f"ERROR: Could not load minor data: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
    
    def _add_electives_sheet(self, writer, semester):
        """Add Electives sheet to the workbook."""
        try:
            electives_df = self._get_electives_data(semester)
            
            if electives_df.empty:
                # Create empty sheet with headers
                electives_df = pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
                print(f"  No elective data found for semester {semester} - creating empty Electives sheet")
            
            # Write to Excel
            electives_df.to_excel(writer, sheet_name='Electives', index=False)
            
            # Format the sheet
            try:
                ws = writer.sheets['Electives']
                self._format_worksheet(ws, has_index=False, start_row=1)
                print(f"  SUCCESS: Added Electives sheet ({len(electives_df)} courses)")
            except Exception as e:
                print(f"  WARNING: Could not format Electives sheet: {e}")
                
        except Exception as e:
            print(f"  WARNING: Could not add Electives sheet: {e}")
    
    def _add_minor_sheet(self, writer, semester):
        """Add Minor sheet to the workbook."""
        try:
            minor_df = self._get_minor_data(semester)
            
            if minor_df.empty:
                # Create empty sheet with headers
                minor_df = pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
                print(f"  No minor data found for semester {semester} - creating empty Minor sheet")
            
            # Write to Excel
            minor_df.to_excel(writer, sheet_name='Minor', index=False)
            
            # Format the sheet
            try:
                ws = writer.sheets['Minor']
                self._format_worksheet(ws, has_index=False, start_row=1)
                print(f"  SUCCESS: Added Minor sheet ({len(minor_df)} courses)")
            except Exception as e:
                print(f"  WARNING: Could not format Minor sheet: {e}")
                
        except Exception as e:
            print(f"  WARNING: Could not add Minor sheet: {e}")
    
    def export_semester7_timetable(self):
        """Export special unified timetable for 7th semester with baskets.
        Creates:
        1. Main timetable showing baskets (7B1, 7B2, 7B3, 7B4) - 9:00 AM to 5:30 PM only, 2 classes per basket
        2. Basket assignments sheet (which courses go to which baskets)"""
        semester = 7
        print(f"\n{'='*60}")
        print(f"GENERATING SEMESTER {semester} UNIFIED TIMETABLE (BASKETS)")
        print(f"{'='*60}")
        # Reset color map for each workbook
        self._course_color_map = {}
        
        filename = f"sem{semester}_timetable.xlsx"
        filepath = FileManager.get_output_path(filename)
        
        # Attempt to open writer
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except PermissionError as pe:
            print(f"\nWARNING: Cannot write to {filepath} (Permission denied / file may be open).")
            import time
            timestamp = int(time.time())
            alt_filename = f"sem{semester}_timetable_{timestamp}.xlsx"
            alt_filepath = FileManager.get_output_path(alt_filename)
            print(f"Attempting alternative filename: {alt_filename}")
            try:
                writer = pd.ExcelWriter(alt_filepath, engine='openpyxl')
                filepath = alt_filepath
                filename = alt_filename
            except Exception as e:
                print(f"\nFAILED: Could not create {filename}: {e}")
                import traceback
                traceback.print_exc()
                return False
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        try:
            with writer as w:
                print(f"Creating {filename}...")
                
                # Get 7th semester courses
                if 'course' not in self.dfs:
                    print("ERROR: Course data not found")
                    return False
                
                course_df = self.dfs['course']
                if 'Semester' not in course_df.columns:
                    print("ERROR: Semester column not found")
                    return False
                
                # Filter 7th semester courses
                temp_df = course_df.copy()
                temp_df['Semester'] = pd.to_numeric(temp_df['Semester'], errors='coerce')
                sem7_courses = temp_df[temp_df['Semester'] == semester].copy()
                
                # Separate baskets and non-basket courses
                if 'Course Code' not in sem7_courses.columns:
                    print("ERROR: Course Code column not found")
                    return False
                
                # Identify baskets (pattern: 7B1, 7B2, 7B3, 7B4, etc.)
                basket_mask = sem7_courses['Course Code'].astype(str).str.match(r'^7B\d+', na=False)
                baskets = sem7_courses[basket_mask].copy()
                non_basket_courses = sem7_courses[~basket_mask].copy()
                
                print(f"Found {len(baskets)} baskets: {', '.join(baskets['Course Code'].astype(str).tolist()) if not baskets.empty else 'None'}")
                print(f"Found {len(non_basket_courses)} non-basket courses")
                
                # 1. Generate unified timetable with baskets
                # For 7th semester: classes only from 9:00 AM to 5:30 PM
                from config import DAYS, TEACHING_SLOTS, LECTURE_DURATION
                from config import LUNCH_SLOTS
                
                # Filter slots to only include 9:00 AM to 5:30 PM (17:30)
                # Slots start from '09:00-09:30' and go until '17:00-17:30'
                sem7_slots = [s for s in TEACHING_SLOTS if s >= '09:00-09:30' and s <= '17:00-17:30']
                
                schedule = pd.DataFrame(index=DAYS, columns=sem7_slots)
                for day in DAYS:
                    for slot in sem7_slots:
                        schedule.loc[day, slot] = 'Free'
                
                # Mark lunch slots
                for day in DAYS:
                    for lunch_slot in LUNCH_SLOTS:
                        if lunch_slot in schedule.columns:
                            schedule.loc[day, lunch_slot] = 'LUNCH BREAK'
                
                # Schedule baskets - assign each basket to different time slots
                # Each basket gets 2 lectures per week (for 7th semester)
                import random
                basket_codes = baskets['Course Code'].astype(str).tolist() if not baskets.empty else []
                
                # Schedule each basket with 2 lectures per week
                for basket_code in basket_codes:
                    scheduled = 0
                    attempts = 0
                    max_attempts = 100
                    
                    while scheduled < 2 and attempts < max_attempts:
                        attempts += 1
                        day = random.choice(DAYS)
                        # Avoid lunch slots - use only sem7_slots
                        available_slots = [s for s in sem7_slots if s not in LUNCH_SLOTS]
                        if not available_slots:
                            continue
                        
                        start_slot = random.choice(available_slots)
                        try:
                            start_idx = sem7_slots.index(start_slot)
                            end_idx = start_idx + LECTURE_DURATION
                            if end_idx > len(sem7_slots):
                                continue
                            slots = sem7_slots[start_idx:end_idx]
                            
                            # Check if all slots are free
                            if all(schedule.loc[day, s] == 'Free' for s in slots):
                                # Check if any slot is lunch
                                if any(s in LUNCH_SLOTS for s in slots):
                                    continue
                                
                                # Assign basket to these slots
                                for slot in slots:
                                    schedule.loc[day, slot] = basket_code
                                scheduled += 1
                        except (ValueError, IndexError):
                            continue
                
                # Write main timetable
                clean_schedule = schedule.replace('Free', '-')
                clean_schedule.to_excel(w, sheet_name='Timetable', index=True, startrow=0)
                
                # Apply color coding
                try:
                    ws = w.sheets['Timetable']
                    self._apply_color_coding(ws, clean_schedule, start_row=1, start_col=1)
                except Exception as e:
                    print(f"    WARNING: Could not apply color coding: {e}")
                
                # Format timetable worksheet
                try:
                    ws = w.sheets['Timetable']
                    self._format_worksheet(ws, has_index=True, start_row=1)
                except Exception as e:
                    print(f"    WARNING: Could not format Timetable: {e}")
                
                print(f"    SUCCESS: Main timetable created with {len(basket_codes)} baskets")
                
                # 2. Create basket assignments sheet
                # Check if there's a "7th sem " sheet with basket assignments
                basket_assignments = pd.DataFrame(columns=['Basket Code', 'Course Code', 'Course Name', 'Department', 'LTPSC', 'Credits', 'Instructor', 'Classroom Allocated'])
                
                # Get non-lab classrooms for basket allocation
                nonlab_rooms = []
                if self.schedule_gen and hasattr(self.schedule_gen, 'nonlab_rooms'):
                    nonlab_rooms = [(name, cap) for name, cap in self.schedule_gen.nonlab_rooms if name and name.upper() != 'C004']
                
                # If no nonlab_rooms, try to get from classrooms (excluding labs)
                if not nonlab_rooms and self.schedule_gen and hasattr(self.schedule_gen, 'classrooms'):
                    lab_room_names = {room[0] for room in getattr(self.schedule_gen, 'lab_rooms', [])}
                    nonlab_rooms = [(name, cap) for name, cap in self.schedule_gen.classrooms 
                                  if name and name not in lab_room_names and name.upper() != 'C004']
                
                # Track room allocation per course within each basket
                # Key: (basket_code, course_code), Value: room_name
                course_room_map = {}
                # Track rooms used per basket to ensure different rooms for courses in same basket
                # Key: basket_code, Value: set of room_names already used
                basket_used_rooms = {}
                # Track global room index for sequential allocation
                room_index = 0
                
                # Look for 7th sem sheet in data_frames
                sem7_sheet_key = None
                for key in self.dfs.keys():
                    key_lower = key.lower()
                    # Match patterns like "course_7th_sem", "7th_sem", etc.
                    if ('7th' in key_lower and 'sem' in key_lower) or key_lower == '7th_sem_':
                        sem7_sheet_key = key
                        break
                
                if sem7_sheet_key and sem7_sheet_key in self.dfs:
                    sem7_sheet_df = self.dfs[sem7_sheet_key]
                    print(f"    Found 7th semester sheet: {sem7_sheet_key} with {len(sem7_sheet_df)} courses")
                    
                    # Map columns from the sheet to our format
                    basket_col = None
                    course_code_col = None
                    course_name_col = None
                    faculty_col = None
                    
                    for col in sem7_sheet_df.columns:
                        col_lower = str(col).lower()
                        if 'basket' in col_lower:
                            basket_col = col
                        elif 'course code' in col_lower:
                            course_code_col = col
                        elif col_lower == 'course' or 'course name' in col_lower:
                            course_name_col = col
                        elif 'faculty' in col_lower or 'instructor' in col_lower:
                            faculty_col = col
                    
                    if basket_col and course_code_col:
                        # Build basket assignments dataframe
                        for _, row in sem7_sheet_df.iterrows():
                            basket_code = str(row.get(basket_col, '')).strip()
                            course_code = str(row.get(course_code_col, '')).strip()
                            course_name = str(row.get(course_name_col, '')).strip() if course_name_col else ''
                            instructor = str(row.get(faculty_col, '')).strip() if faculty_col else ''
                            
                            # Allocate classroom for this course (different room for each course in same basket)
                            allocated_room = ''
                            course_key = (basket_code, course_code)
                            
                            if basket_code and course_code:
                                # Check if this course already has a room assigned
                                if course_key in course_room_map:
                                    allocated_room = course_room_map[course_key]
                                else:
                                    # Get rooms already used for other courses in this basket
                                    used_rooms = basket_used_rooms.get(basket_code, set())
                                    
                                    # Find next available room not used in this basket
                                    found_room = False
                                    start_index = room_index
                                    
                                    # Try to find a room not used in this basket
                                    for _ in range(len(nonlab_rooms)):
                                        if nonlab_rooms and room_index < len(nonlab_rooms):
                                            candidate_room = nonlab_rooms[room_index][0]
                                            if candidate_room not in used_rooms:
                                                allocated_room = candidate_room
                                                course_room_map[course_key] = allocated_room
                                                basket_used_rooms.setdefault(basket_code, set()).add(allocated_room)
                                                room_index = (room_index + 1) % len(nonlab_rooms)
                                                found_room = True
                                                break
                                            room_index = (room_index + 1) % len(nonlab_rooms)
                                        else:
                                            room_index = 0
                                    
                                    # If all rooms are used in this basket, cycle through all available
                                    if not found_room and nonlab_rooms:
                                        for i in range(len(nonlab_rooms)):
                                            idx = (start_index + i) % len(nonlab_rooms)
                                            candidate_room = nonlab_rooms[idx][0]
                                            if candidate_room not in used_rooms or len(used_rooms) >= len(nonlab_rooms):
                                                allocated_room = candidate_room
                                                course_room_map[course_key] = allocated_room
                                                basket_used_rooms.setdefault(basket_code, set()).add(allocated_room)
                                                room_index = (idx + 1) % len(nonlab_rooms)
                                                break
                            
                            # Try to get additional info from main course data if available
                            dept = ''
                            ltpsc = ''
                            credits = ''
                            
                            if not course_df.empty and 'Course Code' in course_df.columns:
                                course_match = course_df[course_df['Course Code'].astype(str) == course_code]
                                if not course_match.empty:
                                    match_row = course_match.iloc[0]
                                    dept = str(match_row.get('Department', '')) if 'Department' in match_row else ''
                                    ltpsc = str(match_row.get('LTPSC', '')) if 'LTPSC' in match_row else ''
                                    credits = str(match_row.get('Credits', '')) if 'Credits' in match_row else ''
                            
                            basket_assignments = pd.concat([
                                basket_assignments,
                                pd.DataFrame([{
                                    'Basket Code': basket_code,
                                    'Course Code': course_code,
                                    'Course Name': course_name,
                                    'Department': dept,
                                    'LTPSC': ltpsc,
                                    'Credits': credits,
                                    'Instructor': instructor,
                                    'Classroom Allocated': allocated_room
                                }])
                            ], ignore_index=True)
                else:
                    # No 7th sem sheet found - create empty rows for each basket
                    if not baskets.empty:
                        for _, basket_row in baskets.iterrows():
                            basket_code = str(basket_row.get('Course Code', ''))
                            course_code = ''  # Empty course code for basket-only rows
                            
                            # Allocate classroom for this basket
                            allocated_room = ''
                            course_key = (basket_code, course_code)
                            
                            if basket_code:
                                if course_key in course_room_map:
                                    allocated_room = course_room_map[course_key]
                                else:
                                    # Get rooms already used for this basket
                                    used_rooms = basket_used_rooms.get(basket_code, set())
                                    
                                    # Find next available room not used in this basket
                                    found_room = False
                                    for _ in range(len(nonlab_rooms)):
                                        if nonlab_rooms and room_index < len(nonlab_rooms):
                                            candidate_room = nonlab_rooms[room_index][0]
                                            if candidate_room not in used_rooms:
                                                allocated_room = candidate_room
                                                course_room_map[course_key] = allocated_room
                                                basket_used_rooms.setdefault(basket_code, set()).add(allocated_room)
                                                room_index = (room_index + 1) % len(nonlab_rooms)
                                                found_room = True
                                                break
                                            room_index = (room_index + 1) % len(nonlab_rooms)
                                        else:
                                            room_index = 0
                                    
                                    if not found_room and nonlab_rooms:
                                        allocated_room = nonlab_rooms[0][0] if nonlab_rooms else ''
                                        course_room_map[course_key] = allocated_room
                                        basket_used_rooms.setdefault(basket_code, set()).add(allocated_room)
                            
                            basket_assignments = pd.concat([
                                basket_assignments,
                                pd.DataFrame([{
                                    'Basket Code': basket_code,
                                    'Course Code': '',
                                    'Course Name': '',
                                    'Department': '',
                                    'LTPSC': '',
                                    'Credits': '',
                                    'Instructor': '',
                                    'Classroom Allocated': allocated_room
                                }])
                            ], ignore_index=True)
                
                total_courses = len(course_room_map)
                total_rooms_used = len(set(course_room_map.values()))
                if course_room_map:
                    print(f"    Allocated {total_rooms_used} unique classrooms to {total_courses} courses across baskets")
                
                basket_assignments.to_excel(w, sheet_name='Basket_Assignments', index=False)
                
                # Format basket assignments worksheet
                try:
                    ws = w.sheets['Basket_Assignments']
                    self._format_worksheet(ws, has_index=False, start_row=1)
                except Exception as e:
                    print(f"    WARNING: Could not format Basket_Assignments: {e}")
                
                print(f"    SUCCESS: Basket assignments sheet created with {len(basket_assignments)} entries")
                
                print(f"\nSUCCESS: Created {filename}")
                print(f"  - Unified timetable with baskets (9:00 AM - 5:30 PM, 2 classes per basket)")
                print(f"  - Basket assignments sheet")
            
            return True
            
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False